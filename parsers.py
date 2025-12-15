import os
import xlrd
import openpyxl
from datetime import datetime
# Import addr_to_index yang baru dibuat
from helpers import clean_currency, detect_currency_from_text, addr_to_index

# ==========================================
# 1. ABSTRAKSI (ADAPTER PATTERN)
# ==========================================

class ExcelAdapter:
    """Kelas induk Adapter."""
    
    def get_val(self, row, col):
        raise NotImplementedError
    
    def get_by_addr(self, addr):
        """Helper: Ambil nilai berdasarkan alamat string 'K4'"""
        r, c = addr_to_index(addr)
        if r is None or c is None: return None
        return self.get_val(r, c)

    def get_date_tuple(self, row, col):
        raise NotImplementedError

    def get_date_by_addr(self, addr):
        """Helper: Ambil tanggal berdasarkan alamat string 'B3'"""
        r, c = addr_to_index(addr)
        if r is None or c is None: return "", datetime.min
        return self.get_date_tuple(r, c)

    @property
    def max_rows(self):
        raise NotImplementedError

class XlrdAdapter(ExcelAdapter):
    def __init__(self, sheet, datemode):
        self.sheet = sheet
        self.datemode = datemode

    def get_val(self, row, col):
        try:
            return self.sheet.cell_value(row, col)
        except:
            return None

    def get_date_tuple(self, row, col):
        val = self.get_val(row, col)
        if isinstance(val, float):
            try:
                dt_tuple = xlrd.xldate_as_tuple(val, self.datemode)
                dt_obj = datetime(*dt_tuple)
                return dt_obj.strftime("%d-%b-%y"), dt_obj
            except: pass
        return str(val), datetime.min

    @property
    def max_rows(self):
        return self.sheet.nrows

class OpenpyxlAdapter(ExcelAdapter):
    def __init__(self, sheet):
        self.sheet = sheet

    def get_val(self, row, col):
        try:
            # Openpyxl 1-based index
            return self.sheet.cell(row=row+1, column=col+1).value
        except:
            return None

    def get_date_tuple(self, row, col):
        val = self.get_val(row, col)
        if isinstance(val, datetime):
            return val.strftime("%d-%b-%y"), val
        return str(val) if val else "", datetime.min

    @property
    def max_rows(self):
        return self.sheet.max_row

# ==========================================
# 2. LOGIKA BISNIS (CORE PARSER)
# ==========================================

def extract_common_logic(adapter: ExcelAdapter):
    try:
        # 1. Ambil Tanggal
        date_str, date_obj = adapter.get_date_by_addr("B3")

        # 2. Deteksi Currency
        raw_a5 = adapter.get_by_addr("A5")
        detected_ccy = detect_currency_from_text(raw_a5)

        # 3. Scanning Baris
        sub_total = 0; penalty = 0; warranty = 0; total_cost = 0; cm_booked = 0; cr_booked = 0
        limit = min(adapter.max_rows, 150)

        for r in range(9, limit):
            raw = adapter.get_val(r, 0)
            txt = str(raw).upper() if raw else ""
            if not txt: continue
            val_col = 4 
            
            if "SUB TOTAL" in txt and sub_total == 0: 
                sub_total = adapter.get_val(r, val_col)
            elif "PENALTY" in txt and penalty == 0: 
                penalty = adapter.get_val(r, val_col)
            elif "WARRANTY" in txt and warranty == 0: 
                warranty = adapter.get_val(r, val_col)
            elif "WARRANTTY" in txt and warranty == 0: 
                warranty = adapter.get_val(r, val_col)
            elif "TOTAL COST" in txt and total_cost == 0: 
                total_cost = adapter.get_val(r, val_col)
            elif "CM BOOKED" in txt and cm_booked == 0: 
                cm_booked = adapter.get_val(r, val_col)
            elif "CR BOOKED" in txt and cr_booked == 0: 
                cr_booked = adapter.get_val(r, val_col)

        # 4. Header Info (Pencarian Dinamis)
        project_no = None
        cust_name = None
        
        found = False
        for r in range(11): 
            if found: break
            for c in range(21):
                val = adapter.get_val(r, c)
                if val and str(val).strip().upper().startswith("PROJECT NO"):
                    project_no = adapter.get_val(r, c + 1)
                    if r > 0: cust_name = adapter.get_val(r - 1, c + 1)
                    found = True
                    break
        
        if not project_no:
            project_no = adapter.get_by_addr("K4")
            cust_name = adapter.get_by_addr("K3")
        if not project_no:
            project_no = adapter.get_by_addr("H4")
            cust_name = adapter.get_by_addr("H3")

        # 5. Ambil Nilai Lainnya
        kurs = clean_currency(adapter.get_by_addr("B4"))
        project_val = clean_currency(adapter.get_by_addr("B5"))
        
        # --- VALIDASI KELENGKAPAN DATA ---
        status = "OK"
        msg = ""

        # Cek Project Value
        if not project_val or project_val == 0:
            status = "DATA INCOMPLETE"
            msg = "Project Value 0/Kosong"
        
        # Cek Sub Total (Indikator parsing baris gagal/data kosong)
        elif not sub_total or sub_total == 0:
            status = "DATA INCOMPLETE"
            msg = "Sub Total Kosong/Gagal Parse"
            
        # Cek Tanggal
        elif not date_str:
            status = "DATA INCOMPLETE"
            msg = "Tanggal Proyek Kosong"

        # 6. Return Data
        return {
            "status": status,
            "msg": msg, # Pesan error jika ada
            "_sort_date": date_obj,
            "Project No": project_no,
            "Cust Name": cust_name,
            "Proj Date": date_str,
            "Currency": detected_ccy,
            "Kurs": kurs,
            "Project Value": project_val,
            "Sub Total": clean_currency(sub_total),
            "Penalty": clean_currency(penalty),
            "Warranty": clean_currency(warranty),
            "Total Cost": clean_currency(total_cost),
            "CM Booked": clean_currency(cm_booked),
            "CR Booked": clean_currency(cr_booked)
        }

    except Exception as e:
        return {"status": "ERROR", "msg": str(e), "_sort_date": datetime.min}

# ==========================================
# 3. ENTRY POINTS
# ==========================================

def parse_xls_classic(filepath):
    try:
        wb = xlrd.open_workbook(filepath, formatting_info=False)
        sheet = wb.sheet_by_index(0)
        adapter = XlrdAdapter(sheet, wb.datemode)
        return extract_common_logic(adapter)
    except Exception as e:
        return {"status": "ERROR", "msg": f"XLS Error: {str(e)}", "_sort_date": datetime.min}

def parse_xlsx_modern(filepath):
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        sheet = wb.active
        adapter = OpenpyxlAdapter(sheet)
        return extract_common_logic(adapter)
    except Exception as e:
        return {"status": "ERROR", "msg": f"XLSX Error: {str(e)}", "_sort_date": datetime.min}

def extract_dispatcher(filepath):
    ext = os.path.splitext(filepath)[1].lower()
    
    if ext == ".xls":
        data = parse_xls_classic(filepath)
    elif ext == ".xlsx":
        data = parse_xlsx_modern(filepath)
    else:
        return {"status": "SKIP", "msg": "Format tidak didukung", "_sort_date": datetime.min}
    
    if data["status"] == "OK":
        if not data.get("Project No"):
            data["status"] = "PARSING ERROR"
            data["msg"] = "Project No Kosong"
    
    return data