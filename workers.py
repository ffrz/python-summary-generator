import os
import shutil
import openpyxl
from datetime import datetime
from PySide6.QtCore import QThread, Signal
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler

from helpers import sanitize_filename, extract_year_from_date
from parsers import extract_dispatcher

# ==========================================
# WATCHER THREAD (MONITORING)
# ==========================================

class FolderChangeHandler(FileSystemEventHandler):
    def __init__(self, signal_emitter):
        self.signal_emitter = signal_emitter

    def on_any_event(self, event):
        if event.is_directory: return
        filename = os.path.basename(event.src_path)
        if filename.startswith("~$"): return
        if filename.lower().endswith(('.xls', '.xlsx')):
            self.signal_emitter.emit()

class WatcherThread(QThread):
    folder_changed = Signal()

    def __init__(self, folder_path):
        super().__init__()
        self.folder_path = folder_path
        self.observer = None

    def run(self):
        self.observer = Observer()
        event_handler = FolderChangeHandler(self.folder_changed)
        try:
            self.observer.schedule(event_handler, self.folder_path, recursive=False)
            self.observer.start()
            while not self.isInterruptionRequested():
                self.msleep(500)
        except:
            pass
        finally:
            if self.observer:
                self.observer.stop()
                self.observer.join()

    def stop(self):
        self.requestInterruption()

# ==========================================
# WORKER THREADS (SCANNER & GENERATOR)
# ==========================================

class PreviewWorker(QThread):
    progress = Signal(int)
    finished = Signal(list)
    
    def __init__(self, folder_path):
        super().__init__()
        self.folder_path = folder_path
        
    def run(self):
        try:
            all_files = [f for f in os.listdir(self.folder_path) if f.lower().endswith(('.xls', '.xlsx'))]
            all_files = [f for f in all_files if not f.startswith("~$")]
        except:
            all_files = []
        
        total = len(all_files)
        results = []
        
        # 1. PARSE
        for i, filename in enumerate(all_files):
            path = os.path.join(self.folder_path, filename)
            data = extract_dispatcher(path)
            data["filename"] = filename
            data["path"] = path
            results.append(data)
            if total > 0: self.progress.emit(int((i+1)/total * 100))

        # 2. LOGIKA DUPLIKAT
        id_counts = {}
        for item in results:
            if item["status"] == "OK":
                pid = str(item.get("Project No", "")).strip()
                if pid:
                    id_counts[pid] = id_counts.get(pid, 0) + 1
        
        for item in results:
            if item["status"] == "OK":
                pid = str(item.get("Project No", "")).strip()
                if pid and id_counts.get(pid, 0) > 1:
                    item["status"] = "DUPLIKAT" 
        
        # 3. SORTING BY DATE (DEFAULT)
        results.sort(key=lambda x: x.get("_sort_date", datetime.min))
        self.finished.emit(results)

class GeneratorWorker(QThread):
    log_msg = Signal(str)
    finished = Signal(str)
    
    def __init__(self, data_list, output_folder):
        super().__init__()
        self.data_list = data_list
        self.output_folder = output_folder
        
    def run(self):
        self.log_msg.emit("🚀 Memulai proses generate...")
        raw_valid_data = [d for d in self.data_list if d["status"] in ["OK", "DUPLIKAT"]]
        valid_data = raw_valid_data[:10]

        copied_count = 0
        created_paths = set()
        
        # 1. COPY & RENAME
        for item in valid_data:
            try:
                old_path = item["path"]
                ext = os.path.splitext(old_path)[1]
                p_id = sanitize_filename(item['Project No'])
                cust = sanitize_filename(item['Cust Name'])
                year = extract_year_from_date(item.get('Proj Date'))
                
                base_name = f"PCM {p_id} {year} {cust}"
                new_name = f"{base_name}{ext}"
                new_path = os.path.join(self.output_folder, new_name)
                
                counter = 1
                while new_path in created_paths or (os.path.exists(new_path) and new_path not in created_paths):
                    if new_path not in created_paths: break 
                    new_name = f"{base_name} ({counter}){ext}"
                    new_path = os.path.join(self.output_folder, new_name)
                    counter += 1

                created_paths.add(new_path)
                shutil.copy2(old_path, new_path)
                copied_count += 1
                
            except Exception as e:
                self.log_msg.emit(f"❌ Gagal copy {item['filename']}: {e}")

        self.log_msg.emit(f"✅ Berhasil menyalin {copied_count} file.")

        # 2. GENERATE SUMMARY EXCEL
        try:
            current_year = datetime.now().year

            self.log_msg.emit("📊 Membuat file summary...")
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"PCM {current_year} SUMMARY"
            
            # --- A. SETUP JUDUL (Row 1) ---
            # Judul digeser sedikit agar rapi (misal di kolom C)
            ws['B1'] = f"PCM {current_year} SUMMARY" 
            ws['B1'].font = openpyxl.styles.Font(size=14, bold=True, name='Calibri')
            
            # --- B. SETUP HEADER (Row 3) ---
            # Total 21 Kolom (1 Kolom Baru + 20 Kolom Lama)
            headers = [
                "File name",      # Col 1 (A)
                "No",             # Col 2 (B)
                "Project no.",    # Col 3 (C)
                "Busunit",        # Col 4 (D)
                "Proj date",      # Col 5 (E)
                "Cust name",      # Col 6 (F)
                "Ccy",            # Col 7 (G)
                "Project value",  # Col 8 (H)
                "Kurs",           # Col 9 (I)
                "Proj IDR",       # Col 10 (J)
                "BARANG&JASA",    # Col 11 (K)
                "Penalty",        # Col 12 (L)
                "Warranty",       # Col 13 (M)
                "Freight",        # Col 14 (N)
                "Cost (estd.)",   # Col 15 (O)
                "CM booked",      # Col 16 (P)
                "CR booked",      # Col 17 (Q)
                "CM IDR",         # Col 18 (R)
                "CM %",           # Col 19 (S)
                "COST %",         # Col 20 (T)
                "Ket."            # Col 21 (U)
            ]
            
            header_font = openpyxl.styles.Font(bold=True, name='Calibri', size=11)
            header_fill = openpyxl.styles.PatternFill("solid", fgColor="00FFFF") 
            
            black_side = openpyxl.styles.Side(style='thin', color="000000")
            border_black = openpyxl.styles.Border(left=black_side, right=black_side, top=black_side, bottom=black_side)
            border_black_row = openpyxl.styles.Border(left=black_side, right=black_side)
            
            duplicate_fill = openpyxl.styles.PatternFill("solid", fgColor="FFFF00") 
            
            header_row_idx = 3
            ws.append([]) # Row 2 Kosong
            ws.append(headers) # Row 3 Header
            
            for col_num, cell in enumerate(ws[header_row_idx], 1):
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border_black
                cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

            # --- C. ISI DATA (Mulai Row 4) ---
            start_data_row = header_row_idx + 1 
            end_data_row = start_data_row + len(valid_data) - 1

            for idx, item in enumerate(valid_data, 1):
                r = header_row_idx + idx # Row index di Excel
                
                # --- LOGIKA DATA ---
                val_project = item["Project Value"]
                val_ccy = item.get("Currency", "IDR")
                
                raw_kurs = item.get("Kurs", 1.0)
                val_kurs = 1.0 if val_ccy == "IDR" else (raw_kurs if raw_kurs else 1.0)
                
                val_cm = item["CM Booked"]
                
                # --- FIX DATE ---
                val_date = item.get("_sort_date", datetime.min)
                if val_date == datetime.min:
                    val_date = item.get("Proj Date", "")

                # --- RUMUS EXCEL (GESER KOLOM +1 dari sebelumnya) ---
                # H = Col 8 (Project Value), I = Col 9 (Kurs)
                f_proj_idr = f"=H{r}*I{r}" 
                
                # K=11, L=12, M=13, N=14 -> Cost Estd = SUM(K:N)
                val_cost = f"=SUM(K{r}:N{r})" 
                
                f_cr_booked = item["CR Booked"]
                
                # J = 10 (Proj IDR), O = 15 (Cost), R = 18 (CM IDR)
                # CM IDR = Proj IDR - Cost ?? atau CM Booked? -> Sesuai kode lama: CM IDR = J - O
                # TAPI di kode lama f_cm_idr = I - N (Proj IDR - Cost).
                # Sekarang Proj IDR = J, Cost = O. Jadi:
                f_cm_idr = f"=J{r}-O{r}" 
                
                # CM % = CM IDR / Proj IDR -> R / J
                f_cm_pct = f"=IF(J{r}=0, 0, R{r}/J{r})"
                
                # COST % = Cost / Proj IDR -> O / J
                f_cost_pct = f"=IF(S{r}=0, 0, 1-S{r})"
                
                status_ket = ""
                if item["status"] == "DUPLIKAT":
                    status_ket = "Duplikat Input"

                # Mapping Data (Perhatikan urutan)
                row_data = [
                    item["filename"],   # 1. Nama File Asli
                    idx,                # 2. No
                    item["Project No"], # 3. Project No
                    "",                 # 4. Busunit
                    val_date,           # 5. Proj Date
                    item["Cust Name"],  # 6. Cust Name
                    val_ccy,            # 7. Ccy
                    val_project,        # 8. Project Value
                    val_kurs,           # 9. Kurs
                    f_proj_idr,         # 10. Proj IDR
                    item["Sub Total"],  # 11. B&J
                    item["Penalty"],    # 12. Penalty
                    item["Warranty"],   # 13. Warranty
                    0,                  # 14. Freight
                    val_cost,           # 15. Cost Estd
                    val_cm,             # 16. CM Booked
                    f_cr_booked,        # 17. CR Booked
                    f_cm_idr,           # 18. CM IDR
                    f_cm_pct,           # 19. CM %
                    f_cost_pct,         # 20. Cost %
                    status_ket          # 21. Ket
                ]
                
                ws.append(row_data)
                
                # Styling Baris
                for c, val in enumerate(row_data, 1):
                    cell = ws.cell(row=r, column=c)
                    cell.border = border_black_row 
                    
                    # Format Tanggal (Kolom 5)
                    if c == 5:
                        cell.number_format = 'd-mmm-yy'

                    # Format Ribuan (Kolom 8, 10-16, 18)
                    if c in [8, 10, 11, 12, 13, 14, 15, 16, 18]: 
                        cell.number_format = '#,##0'
                    
                    # Format Kurs (Kolom 9)
                    if c == 9:
                        cell.number_format = '#,##0.00'
                    
                    # Format Persen (Kolom 17, 19, 20)
                    if c in [17, 19, 20]:
                        cell.number_format = '0.00%'
                        
                    if item["status"] == "DUPLIKAT":
                        cell.fill = duplicate_fill

            # --- D. TAMBAHKAN BARIS TOTAL (SUMMARY) ---
            if valid_data:
                r_total = end_data_row + 1
                
                total_font = openpyxl.styles.Font(bold=True, name='Calibri', size=11)
                total_border = openpyxl.styles.Border(top=black_side, bottom=openpyxl.styles.Side(style='medium', color="000000"))
                
                # Label GRAND TOTAL di kolom Ccy (Kolom 7)
                ws.cell(row=r_total, column=7, value="GRAND TOTAL").font = total_font
                
                # Kolom yang akan di-SUM (Indeks Bergeser +1 dari kode lama)
                # Lama: [7, 9, 10, 11, 12, 13, 14, 15, 17]
                # Baru: [8, 10, 11, 12, 13, 14, 15, 16, 18]
                sum_cols = [8, 10, 11, 12, 13, 14, 15, 16, 18]
                
                for c in range(1, len(headers) + 1):
                    cell = ws.cell(row=r_total, column=c)
                    
                    if c in sum_cols:
                        col_letter = openpyxl.utils.get_column_letter(c)
                        sum_formula = f"=SUM({col_letter}{start_data_row}:{col_letter}{end_data_row})"
                        cell.value = sum_formula
                        cell.number_format = '#,##0'
                    
                    # Styling Border (Kecuali kolom label Ccy = 7)
                    if c != 7 and c not in sum_cols:
                         cell.border = openpyxl.styles.Border(top=black_side, bottom=openpyxl.styles.Side(style='medium', color="000000"))
                    else:
                         cell.border = total_border
                    
                    cell.font = total_font

            # --- E. FINALISASI ---
            dims = {}
            for row in ws.rows:
                for cell in row:
                    if cell.value:
                        dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))    
            for col, value in dims.items():
                ws.column_dimensions[col].width = value + 2

            summary_name = f"PCM {current_year} SUMMARY.xlsx"
            summary_path = os.path.join(self.output_folder, summary_name)
            
            wb.save(summary_path)
            self.finished.emit(summary_path)

        except Exception as e:
            self.finished.emit(f"ERROR: {e}")