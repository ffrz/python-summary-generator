import sys
import os
import re
import shutil
import xlrd
import openpyxl
from datetime import datetime

# --- LIBRARY WATCHDOG UNTUK MONITORING ---
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler

from PySide6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, 
                               QHBoxLayout, QPushButton, QLabel, QProgressBar, 
                               QTableWidget, QTableWidgetItem, QFileDialog, 
                               QMessageBox, QHeaderView, QAbstractItemView)
from PySide6.QtCore import Qt, QThread, Signal, QSettings, QUrl, QTimer
from PySide6.QtGui import QColor, QDesktopServices, QAction

# ==========================================
# 1. HELPER FUNCTIONS
# ==========================================

def sanitize_filename(name):
    """Membersihkan karakter ilegal untuk nama file"""
    if not name: return "Unknown"
    clean = re.sub(r'[\\/*?:"<>|]', "", str(name)).strip()
    return clean if clean else "Unknown"

def clean_currency(value):
    if value in [None, ""]: return 0
    if isinstance(value, (int, float)): return value
    str_val = str(value).replace("Rp", "").replace(".", "").replace(",", ".").strip()
    try: return float(str_val)
    except: return 0

def extract_year_from_date(date_str):
    try:
        if not date_str: return str(datetime.now().year)
        parts = date_str.split('-')
        if len(parts) == 3:
            yy = parts[2]
            return f"20{yy}" if len(yy) == 2 else yy
    except:
        pass
    return str(datetime.now().year)

# --- PARSER KHUSUS .XLS (XLRD) ---
def parse_xls_classic(filepath):
    try:
        wb = xlrd.open_workbook(filepath, formatting_info=False)
        sheet = wb.sheet_by_index(0)
        
        def get_xls_val(addr):
            match = re.match(r"([A-Z]+)([0-9]+)", addr.upper())
            if not match: return None
            col_str, row_str = match.groups()
            row = int(row_str) - 1
            col = 0
            for char in col_str:
                col = col * 26 + (ord(char) - ord('A') + 1)
            col -= 1
            try: return sheet.cell_value(row, col)
            except: return None

        def get_xls_date_pack(addr):
            val = get_xls_val(addr)
            if isinstance(val, float):
                try:
                    dt_tuple = xlrd.xldate_as_tuple(val, wb.datemode)
                    dt_obj = datetime(*dt_tuple)
                    return dt_obj.strftime("%d-%b-%y"), dt_obj
                except: pass
            return str(val), datetime.min

        date_str, date_obj = get_xls_date_pack("B3")
        sub_total = 0; penalty = 0; warranty = 0; total_cost = 0; cm_booked = 0; cr_booked = 0
        limit = min(sheet.nrows, 150)
        
        for r in range(9, limit):
            try:
                raw = sheet.cell_value(r, 0)
                txt = str(raw).upper() if raw else ""
            except: continue
            
            if "SUB TOTAL" in txt and sub_total == 0: sub_total = sheet.cell_value(r, 4)
            elif "PENALTY" in txt and penalty == 0: penalty = sheet.cell_value(r, 4)
            elif "WARRANTY" in txt and warranty == 0: warranty = sheet.cell_value(r, 4)
            elif "WARRANTTY" in txt and warranty == 0: warranty = sheet.cell_value(r, 4)
            elif "TOTAL COST" in txt and total_cost == 0: total_cost = sheet.cell_value(r, 4)
            elif "CM BOOKED" in txt and cm_booked == 0: cm_booked = sheet.cell_value(r, 4)
            elif "CR BOOKED" in txt and cr_booked == 0: cr_booked = sheet.cell_value(r, 4)

            cust_name = get_xls_val("K3")
            project_no = get_xls_val("K4")
            if not project_no:
                cust_name = get_xls_val("H3")
                project_no = get_xls_val("H4")

        return {
            "status": "OK",
            "_sort_date": date_obj,
            "Project No": project_no,
            "Cust Name": cust_name,
            "Proj Date": date_str,
            "Kurs": clean_currency(get_xls_val("B4")),
            "Project Value": clean_currency(get_xls_val("B5")),
            "Sub Total": clean_currency(sub_total),
            "Penalty": clean_currency(penalty),
            "Warranty": clean_currency(warranty),
            "Total Cost": clean_currency(total_cost),
            "CM Booked": clean_currency(cm_booked),
            "CR Booked": clean_currency(cr_booked)
        }
    except Exception as e:
        return {"status": "ERROR", "msg": str(e), "_sort_date": datetime.min}

# --- PARSER KHUSUS .XLSX (OPENPYXL) ---
def parse_xlsx_modern(filepath):
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        sheet = wb.active
        
        def get_xlsx_val(addr):
            return sheet[addr].value

        def get_xlsx_date_pack(addr):
            val = sheet[addr].value
            if isinstance(val, datetime):
                return val.strftime("%d-%b-%y"), val
            return str(val) if val else "", datetime.min

        date_str, date_obj = get_xlsx_date_pack("B3")
        sub_total = 0; penalty = 0; warranty = 0; total_cost = 0; cm_booked = 0; cr_booked = 0
        limit = min(sheet.max_row, 150)
        
        for r in range(10, limit + 1):
            try:
                cell_val = sheet.cell(row=r, column=1).value
                txt = str(cell_val).upper() if cell_val else ""
            except: continue
            
            val_col = 5
            if "SUB TOTAL" in txt and sub_total == 0: sub_total = sheet.cell(row=r, column=val_col).value
            elif "PENALTY" in txt and penalty == 0: penalty = sheet.cell(row=r, column=val_col).value
            elif "WARRANTY" in txt and warranty == 0: warranty = sheet.cell(row=r, column=val_col).value
            elif "WARRANTTY" in txt and warranty == 0: warranty = sheet.cell(row=r, column=val_col).value
            elif "TOTAL COST" in txt and total_cost == 0: total_cost = sheet.cell(row=r, column=val_col).value
            elif "CM BOOKED" in txt and cm_booked == 0: cm_booked = sheet.cell(row=r, column=val_col).value
            elif "CR BOOKED" in txt and cr_booked == 0: cr_booked = sheet.cell(row=r, column=val_col).value

            cust_name = get_xlsx_val("K3")
            project_no = get_xlsx_val("K4")

            if not project_no:
                cust_name = get_xlsx_val("H3")
                project_no = get_xlsx_val("H4")

        return {
            "status": "OK",
            "_sort_date": date_obj,
            "Project No": project_no,
            "Cust Name": cust_name,
            "Proj Date": date_str,
            "Kurs": clean_currency(get_xlsx_val("B4")),
            "Project Value": clean_currency(get_xlsx_val("B5")),
            "Sub Total": clean_currency(sub_total),
            "Penalty": clean_currency(penalty),
            "Warranty": clean_currency(warranty),
            "Total Cost": clean_currency(total_cost),
            "CM Booked": clean_currency(cm_booked),
            "CR Booked": clean_currency(cr_booked)
        }

    except Exception as e:
        return {"status": "ERROR", "msg": str(e), "_sort_date": datetime.min}

def extract_dispatcher(filepath):
    ext = os.path.splitext(filepath)[1].lower()
    data = {}
    
    if ext == ".xls":
        data = parse_xls_classic(filepath)
    elif ext == ".xlsx":
        data = parse_xlsx_modern(filepath)
    else:
        return {"status": "SKIP", "msg": "Format tidak didukung", "_sort_date": datetime.min}
    
    if data["status"] == "OK":
        if not data.get("Project No"):
            data["status"] = "PARSING ERROR"
            data["msg"] = "Project No Kosong"
    
    return data

# ==========================================
# 2. WATCHER THREAD (MONITORING)
# ==========================================

class FolderChangeHandler(FileSystemEventHandler):
    def __init__(self, signal_emitter):
        self.signal_emitter = signal_emitter

    def on_any_event(self, event):
        if event.is_directory: return
        filename = os.path.basename(event.src_path)
        if filename.startswith("~$"): return
        if filename.lower().endswith(('.xls', '.xlsx')):
            self.signal_emitter.emit()

class WatcherThread(QThread):
    folder_changed = Signal()

    def __init__(self, folder_path):
        super().__init__()
        self.folder_path = folder_path
        self.observer = None

    def run(self):
        self.observer = Observer()
        event_handler = FolderChangeHandler(self.folder_changed)
        try:
            self.observer.schedule(event_handler, self.folder_path, recursive=False)
            self.observer.start()
            while not self.isInterruptionRequested():
                self.msleep(500)
        except:
            pass
        finally:
            if self.observer:
                self.observer.stop()
                self.observer.join()

    def stop(self):
        self.requestInterruption()

# ==========================================
# 3. WORKER THREADS (SCANNER & GENERATOR)
# ==========================================

class PreviewWorker(QThread):
    progress = Signal(int)
    finished = Signal(list)
    
    def __init__(self, folder_path):
        super().__init__()
        self.folder_path = folder_path
        
    def run(self):
        try:
            all_files = [f for f in os.listdir(self.folder_path) if f.lower().endswith(('.xls', '.xlsx'))]
            all_files = [f for f in all_files if not f.startswith("~$")]
        except:
            all_files = []
        
        total = len(all_files)
        results = []
        
        # 1. PARSE
        for i, filename in enumerate(all_files):
            path = os.path.join(self.folder_path, filename)
            data = extract_dispatcher(path)
            data["filename"] = filename
            data["path"] = path
            results.append(data)
            if total > 0: self.progress.emit(int((i+1)/total * 100))

        # 2. LOGIKA DUPLIKAT
        id_counts = {}
        for item in results:
            if item["status"] == "OK":
                pid = str(item.get("Project No", "")).strip()
                if pid:
                    id_counts[pid] = id_counts.get(pid, 0) + 1
        
        for item in results:
            if item["status"] == "OK":
                pid = str(item.get("Project No", "")).strip()
                if pid and id_counts.get(pid, 0) > 1:
                    item["status"] = "DUPLIKAT" 
        
        # 3. SORTING BY DATE (DEFAULT)
        results.sort(key=lambda x: x.get("_sort_date", datetime.min))
        self.finished.emit(results)

class GeneratorWorker(QThread):
    log_msg = Signal(str)
    finished = Signal(str)
    
    def __init__(self, data_list, output_folder):
        super().__init__()
        self.data_list = data_list
        self.output_folder = output_folder
        
    def run(self):
        self.log_msg.emit("🚀 Memulai proses generate...")
        valid_data = [d for d in self.data_list if d["status"] in ["OK", "DUPLIKAT"]]
        copied_count = 0
        created_paths = set()
        
        # 1. COPY & RENAME (Tetap sama seperti kode asli)
        for item in valid_data:
            try:
                old_path = item["path"]
                ext = os.path.splitext(old_path)[1]
                p_id = sanitize_filename(item['Project No'])
                cust = sanitize_filename(item['Cust Name'])
                year = extract_year_from_date(item.get('Proj Date'))
                
                base_name = f"PCM {p_id} {year} {cust}"
                new_name = f"{base_name}{ext}"
                new_path = os.path.join(self.output_folder, new_name)
                
                counter = 1
                while new_path in created_paths or (os.path.exists(new_path) and new_path not in created_paths):
                    if new_path not in created_paths: break 
                    new_name = f"{base_name} ({counter}){ext}"
                    new_path = os.path.join(self.output_folder, new_name)
                    counter += 1

                created_paths.add(new_path)
                shutil.copy2(old_path, new_path)
                copied_count += 1
                
            except Exception as e:
                self.log_msg.emit(f"❌ Gagal copy {item['filename']}: {e}")

        self.log_msg.emit(f"✅ Berhasil menyalin {copied_count} file.")

        # 2. GENERATE SUMMARY EXCEL (DIMODIFIKASI SESUAI GAMBAR)
        try:
            self.log_msg.emit("📊 Membuat file summary...")
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "PCM SUMMARY"
            
            # --- A. SETUP JUDUL (Row 1) ---
            current_year = datetime.now().year
            ws['A1'] = f"PCM {current_year} SUMMARY"
            ws['A1'].font = openpyxl.styles.Font(size=14, bold=True, name='Calibri')
            
            # --- B. SETUP HEADER (Row 3) ---
            # Kolom sesuai gambar:
            # 1.No, 2.Project no, 3.Busunit, 4.Proj date, 5.Cust name, 6.Ccy, 
            # 7.Project value, 8.Kurs, 9.Proj IDR, 10.BARANG&JASA, 11.Penalty, 
            # 12.Warranty, 13.Freight, 14.Cost (estd.), 15.CM booked, 16.CR booked, 
            # 17.CM IDR, 18.CM %, 19.COST %, 20.Ket.
            
            headers = [
                "No", "Project no.", "Busunit", "Proj date", "Cust name", "Ccy",
                "Project value", "Kurs", "Proj IDR", "BARANG&JASA", "Penalty",
                "Warranty", "Freight", "Cost (estd.)", "CM booked", "CR booked",
                "CM IDR", "CM %", "COST %", "Ket."
            ]
            
            # Styling Header sesuai Gambar (Background Cyan, Border Biru Tebal/Tipis)
            header_font = openpyxl.styles.Font(bold=True, name='Calibri', size=11)
            # Warna Cyan muda sesuai gambar
            header_fill = openpyxl.styles.PatternFill("solid", fgColor="00FFFF") 
            
            # Border warna hitam
            black_side = openpyxl.styles.Side(style='thin', color="000000")
            border_black = openpyxl.styles.Border(left=black_side, right=black_side, top=black_side, bottom=black_side)
            border_black_row = openpyxl.styles.Border(left=black_side, right=black_side)
            
            duplicate_fill = openpyxl.styles.PatternFill("solid", fgColor="FFFF00") # Kuning untuk duplikat
            
            # Tulis Header di Baris 3
            header_row_idx = 3
            ws.append([]) # Row 2 Kosong
            ws.append(headers) # Row 3 Header
            
            for col_num, cell in enumerate(ws[header_row_idx], 1):
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border_black
                cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

            # --- C. ISI DATA (Mulai Row 4) ---
            for idx, item in enumerate(valid_data, 1):
                r = header_row_idx + idx # Row index di Excel
                
                # Ambil value dasar
                val_project = item["Project Value"]
                #val_kurs = item["Kurs"] if item["Kurs"] else 1.0
                val_ccy = "IDR" # HARD CODED DULU
                val_kurs = 1.0 # Paksa ke 1 karena currency selalu IDR
                val_cost = f"=SUM(J{r}:M{r})"
                val_cm = item["CM Booked"]
                
                # Rumus Excel
                f_proj_idr = f"=G{r}*H{r}"                
                f_cr_booked = item["CR Booked"]
                f_cm_idr = f"=I{r}-N{r}"
                f_cm_pct = f"=IF(N{r}=0, 0, Q{r}/I{r})"
                f_cost_pct = f"=IF(Q{r}=0, 0, 1-R{r})"
                
                status_ket = ""
                if item["status"] == "DUPLIKAT":
                    status_ket = "Duplikat Input"

                # Mapping Data ke Kolom
                row_data = [
                    idx,                        # 1. No
                    item["Project No"],         # 2. Project no.
                    "",                         # 3. Busunit (Kosong/Manual)
                    item["Proj Date"],          # 4. Proj date
                    item["Cust Name"],          # 5. Cust name
                    val_ccy,                    # 6. Ccy
                    val_project,                # 7. Project value
                    val_kurs,                   # 8. Kurs
                    f_proj_idr,                 # 9. Proj IDR (Rumus)
                    item["Sub Total"],          # 10. BARANG&JASA
                    item["Penalty"],            # 11. Penalty
                    item["Warranty"],           # 12. Warranty
                    0,                          # 13. Freight (Default 0/Manual)
                    val_cost,                   # 14. Cost (estd.)
                    val_cm,                     # 15. CM booked
                    f_cr_booked,                # 16. CR booked (Rumus %)
                    f_cm_idr,                   # 17. CM IDR (Rumus Link)
                    f_cm_pct,                   # 18. CM % (Rumus %)
                    f_cost_pct,                 # 19. COST % (Rumus %)
                    status_ket                  # 20. Ket.
                ]
                
                ws.append(row_data)
                
                # Styling Baris Data
                for c, val in enumerate(row_data, 1):
                    cell = ws.cell(row=r, column=c)
                    cell.border = border_black_row # Terapkan border ke row
                    
                    # Format Ribuan (Project Value, Proj IDR, B&J, Penalty, Warranty, Freight, Cost, CM, CM IDR)
                    if c in [7, 9, 10, 11, 12, 13, 14, 15, 17]: 
                        cell.number_format = '#,##0'
                    
                    # Format Kurs (2 desimal)
                    if c == 8:
                        cell.number_format = '#,##0.00'

                    # Format Persentase (CR booked, CM %, COST %)
                    if c in [16, 18, 19]:
                        cell.number_format = '0.00%'
                        
                    # Highlight row duplikat
                    if item["status"] == "DUPLIKAT":
                        cell.fill = duplicate_fill

            # Auto-adjust column width (sedikit styling agar rapi)
            dims = {}
            for row in ws.rows:
                for cell in row:
                    if cell.value:
                        dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))    
            for col, value in dims.items():
                ws.column_dimensions[col].width = value + 2

            summary_name = f"PCM {current_year} SUMMARY.xlsx"
            summary_path = os.path.join(self.output_folder, summary_name)
            
            wb.save(summary_path)
            self.finished.emit(summary_path)

        except Exception as e:
            self.finished.emit(f"ERROR: {e}")

# ==========================================
# 4. MAIN WINDOW
# ==========================================

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle(" PCM Summary Generator v1.0.0")
        self.resize(1000, 650)
        self.settings = QSettings("FahmiSoft", "PCMGenerator")
        
        central = QWidget()
        self.setCentralWidget(central)
        layout = QVBoxLayout(central)
        
        # --- IO ---
        grp_io = QWidget()
        layout_io = QVBoxLayout(grp_io)
        
        h1 = QHBoxLayout()
        self.lbl_input = QLabel("Input Folder: (Belum Dipilih)")
        self.lbl_input.setStyleSheet("background: #e3f2fd; padding: 5px; border-radius: 4px;")
        btn_input = QPushButton("1. Pilih Folder INPUT")
        btn_input.clicked.connect(self.select_input)
        h1.addWidget(self.lbl_input); h1.addWidget(btn_input)
        
        h2 = QHBoxLayout()
        self.lbl_output = QLabel("Output Folder: (Belum Dipilih)")
        self.lbl_output.setStyleSheet("background: #e8f5e9; padding: 5px; border-radius: 4px;")
        btn_output = QPushButton("2. Pilih Folder OUTPUT")
        btn_output.clicked.connect(self.select_output)
        h2.addWidget(self.lbl_output); h2.addWidget(btn_output)
        
        layout_io.addLayout(h1); layout_io.addLayout(h2)
        layout.addWidget(grp_io)
        
        # --- TABLE ---
        self.table = QTableWidget()
        cols = ["Nama File Asli", "Project ID", "Customer", "Date", "Nilai Project", "Status"]
        self.table.setColumnCount(len(cols))
        self.table.setHorizontalHeaderLabels(cols)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        
        # --- FITUR SORTING ---
        self.table.setSortingEnabled(True)
        
        # --- FITUR DOUBLE CLICK ---
        self.table.cellDoubleClicked.connect(self.on_table_double_click)
        
        layout.addWidget(self.table)
        
        # --- ACTION ---
        h3 = QHBoxLayout()
        self.progress = QProgressBar()
        self.btn_gen = QPushButton("3. GENERATE FILES")
        self.btn_gen.setStyleSheet("QPushButton { background-color: #2196F3; color: white; font-weight: bold; padding: 10px; } QPushButton:disabled { background-color: #ccc; }")
        self.btn_gen.setEnabled(False)
        self.btn_gen.clicked.connect(self.start_generation)
        h3.addWidget(self.progress); h3.addWidget(self.btn_gen)
        layout.addLayout(h3)
        
        self.setup_statusbar()
        self.input_dir = ""; self.output_dir = ""; self.data_cache = []
        self.scan_worker = None; self.gen_worker = None
        self.watcher_thread = None
        
        self.debounce_timer = QTimer()
        self.debounce_timer.setSingleShot(True)
        self.debounce_timer.setInterval(1500)
        self.debounce_timer.timeout.connect(self.run_preview_scan)

        self.load_settings()

    def setup_statusbar(self):
        status_bar = self.statusBar()
        lbl_version = QLabel("PCM Summary Generator v 1.0.0")
        status_bar.addWidget(lbl_version)
        btn_about = QPushButton("About")
        btn_about.setFlat(True)
        btn_about.setStyleSheet("font-weight: bold; color: #555;")
        btn_about.clicked.connect(self.show_about_dialog)
        status_bar.addPermanentWidget(btn_about)

    def load_settings(self):
        last_in = self.settings.value("last_input_dir")
        last_out = self.settings.value("last_output_dir")
        if last_in and os.path.exists(last_in):
            self.input_dir = last_in
            self.lbl_input.setText(last_in)
            self.start_watcher()
            self.run_preview_scan()
        if last_out and os.path.exists(last_out):
            self.output_dir = last_out
            self.lbl_output.setText(last_out)
            self.check_ready()

    def select_input(self):
        start_dir = self.input_dir if self.input_dir else ""
        path = QFileDialog.getExistingDirectory(self, "Pilih Input Folder", start_dir)
        if path:
            self.input_dir = path
            self.lbl_input.setText(path)
            self.settings.setValue("last_input_dir", path)
            self.start_watcher()
            self.run_preview_scan()

    def select_output(self):
        start_dir = self.output_dir if self.output_dir else ""
        path = QFileDialog.getExistingDirectory(self, "Pilih Output Folder", start_dir)
        if path:
            self.output_dir = path
            self.lbl_output.setText(path)
            self.settings.setValue("last_output_dir", path)
            self.check_ready()

    def start_watcher(self):
        if self.watcher_thread: self.watcher_thread.stop(); self.watcher_thread.wait()
        self.watcher_thread = WatcherThread(self.input_dir)
        self.watcher_thread.folder_changed.connect(self.on_folder_change_detected)
        self.watcher_thread.start()

    def on_folder_change_detected(self):
        self.statusBar().showMessage("🔍 Mendeteksi perubahan file... Menunggu...", 2000)
        self.debounce_timer.start()

    def run_preview_scan(self):
        self.table.setSortingEnabled(False)
        self.table.setRowCount(0)
        self.btn_gen.setEnabled(False)
        self.scan_worker = PreviewWorker(self.input_dir)
        self.scan_worker.progress.connect(self.progress.setValue)
        self.scan_worker.finished.connect(self.on_preview_done)
        self.scan_worker.start()
        
    def on_preview_done(self, results):
        self.data_cache = results
        self.table.setRowCount(len(results))
        for r, item in enumerate(results):
            status = item["status"]
            color = QColor(Qt.black)
            bg_color = QColor(Qt.white)
            
            if status == "DUPLIKAT":
                bg_color = QColor("#FFEB3B"); status = "DUPLIKAT (Diproses)"
            elif status != "OK":
                bg_color = QColor("#FFCDD2"); color = QColor(Qt.red)
            
            def make_item(text):
                it = QTableWidgetItem(str(text))
                it.setForeground(color); it.setBackground(bg_color)
                return it

            self.table.setItem(r, 0, make_item(item["filename"]))
            self.table.setItem(r, 1, make_item(item.get("Project No", "-")))
            self.table.setItem(r, 2, make_item(item.get("Cust Name", "-")))
            self.table.setItem(r, 3, make_item(item.get("Proj Date", "-")))
            val = item.get("Project Value", 0)
            fmt_val = f"{val:,.0f}" if isinstance(val, (int, float)) else str(val)
            self.table.setItem(r, 4, make_item(fmt_val))
            self.table.setItem(r, 5, make_item(status))
        
        self.table.setSortingEnabled(True)
        self.check_ready()
        self.statusBar().showMessage(f"Scan selesai. Total {len(results)} file.", 3000)

    # --- FITUR DOUBLE CLICK OPEN EXCEL ---
    def on_table_double_click(self, row, col):
        if row < 0 or row >= len(self.data_cache): return
        
        # Karena tabel mungkin di-sort, kita tidak bisa pakai index 'row' langsung ke self.data_cache
        # Kita harus cari data yang sesuai dengan 'filename' di kolom 0 baris tersebut
        filename_in_table = self.table.item(row, 0).text()
        
        # Cari file di cache
        selected_file = None
        for item in self.data_cache:
            if item["filename"] == filename_in_table:
                selected_file = item
                break
        
        if not selected_file: return

        # Tampilkan Konfirmasi
        reply = QMessageBox.question(self, "Edit File Input", 
                                     f"Apakah Anda ingin memodifikasi file ini?\n\n{selected_file['filename']}",
                                     QMessageBox.Yes | QMessageBox.No)
        
        if reply == QMessageBox.Yes:
            file_path = selected_file["path"]
            if os.path.exists(file_path):
                # Buka file (Excel/Default App)
                success = QDesktopServices.openUrl(QUrl.fromLocalFile(file_path))
                if not success:
                    # Fallback: Buka Folder jika gagal buka file
                    folder = os.path.dirname(file_path)
                    QDesktopServices.openUrl(QUrl.fromLocalFile(folder))
            else:
                QMessageBox.warning(self, "Error", "File tidak ditemukan!")

    def check_ready(self):
        valid_count = sum(1 for d in self.data_cache if d["status"] in ["OK", "DUPLIKAT"])
        is_ready = bool(self.input_dir and self.output_dir and valid_count > 0)
        self.btn_gen.setEnabled(is_ready)
        if is_ready: self.btn_gen.setText(f"3. GENERATE ({valid_count} File Valid)")
        else: self.btn_gen.setText("3. GENERATE (Menunggu Input/Output)")

    def start_generation(self):
        if not self.output_dir: return
        try:
            files_in_output = [f for f in os.listdir(self.output_dir) if not f.startswith('.')]
            if files_in_output:
                msg_box = QMessageBox(self)
                msg_box.setIcon(QMessageBox.Warning)
                msg_box.setWindowTitle("Folder Output Tidak Kosong")
                msg_box.setText(f"Folder Output berisi {len(files_in_output)} file/folder.")
                msg_box.setInformativeText("File lama mungkin akan tertimpa.\nLanjutkan?")
                btn_overwrite = msg_box.addButton("Lanjutkan (Overwrite)", QMessageBox.AcceptRole)
                btn_cancel = msg_box.addButton("Batalkan", QMessageBox.RejectRole)
                msg_box.exec()
                if msg_box.clickedButton() == btn_cancel: return
        except Exception as e:
            QMessageBox.critical(self, "Error", str(e)); return
        
        self.gen_worker = GeneratorWorker(self.data_cache, self.output_dir)
        self.gen_worker.log_msg.connect(lambda s: self.progress.setFormat(s))
        self.gen_worker.finished.connect(self.on_generation_finished)
        self.progress.setValue(0); self.progress.setRange(0, 0)
        self.btn_gen.setEnabled(False)
        self.gen_worker.start()

    def on_generation_finished(self, result_msg):
        self.progress.setRange(0, 100); self.progress.setValue(100); self.progress.setFormat("Selesai")
        self.btn_gen.setEnabled(True); self.btn_gen.setText("GENERATE ULANG")
        
        if "ERROR:" in result_msg:
            QMessageBox.critical(self, "Gagal", result_msg)
        else:
            reply = QMessageBox.question(self, "Sukses", 
                                         f"Generate selesai!\nFile Summary:\n{result_msg}\n\n"
                                         "Buka folder output di Windows Explorer?",
                                         QMessageBox.Yes | QMessageBox.No)
            if reply == QMessageBox.Yes:
                QDesktopServices.openUrl(QUrl.fromLocalFile(self.output_dir))

    def show_about_dialog(self):
        info = (
            "<h3>PCM Summary Generator</h3>"
            "<p>Version: 1.0.0</p>"
            "<p>Aplikasi untuk rekapitulasi otomatis Project Cost Management.</p>"
            "<hr>"
            "<p><b>Developer:</b> Fahmi Fauzi Rahman</p>"
            "<p><b>Contact:</b> 0853-1740-4760</p>"
            "<hr>"
            "<p><b>Credits / Libraries Used:</b></p>"
            "<ul>"
            "<li>Python 3</li>"
            "<li>PySide6 (Qt for Python)</li>"
            "<li>openpyxl (Excel Modern Support)</li>"
            "<li>xlrd (Excel Classic Support)</li>"
            "</ul>"
        )
        QMessageBox.about(self, "About", info)

if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    w = MainWindow()
    w.show()
    sys.exit(app.exec())